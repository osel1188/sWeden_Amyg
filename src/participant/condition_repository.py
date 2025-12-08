import pandas as pd
from pathlib import Path
from typing import Union, Optional
from datetime import datetime
import shutil
import openpyxl  # Added for style-preserving writes

class RepositoryError(Exception):
    """Errors related to Excel data reading or writing."""
    pass

# --- Concern 2: Data Repository (Excel) ---

class ConditionRepository:
    """
    Manages reading from and writing to the participant condition/assignment Excel file.
    """
    def __init__(self, excel_file_path: Union[str, Path]):
        self.excel_file_path = excel_file_path
        # MODIFIED: Updated required columns
        self.required_cols = [
            "ID", "Priority Order", "randomization_number", "sex", "condition",
            "Task1", "Task2", "Task3", "Task4", "Task5", "Task6",
            "EM version", "Stroop version"
        ]
        # "session date" is intentionally omitted from required_cols

    def load_data(self) -> pd.DataFrame:
        """
        Loads the participant data from the Excel file.
        
        :raises RepositoryError: If file not found, read error, or columns missing.
        :return: A pandas DataFrame with participant data.
        """
        try:
            df = pd.read_excel(self.excel_file_path)
        except FileNotFoundError:
            raise RepositoryError(f"Excel file not found at: {self.excel_file_path}")
        except Exception as e:
            raise RepositoryError(f"Error reading Excel file: {e}")
            
        # Validate columns
        missing_cols = [col for col in self.required_cols if col not in df.columns]
        if missing_cols:
            raise RepositoryError(f"Missing required columns in Excel: {', '.join(missing_cols)}")
            
        # Ensure 'sex' column is string type for comparison
        df['sex'] = df['sex'].astype(str)
        return df

    def find_by_id(self, df: pd.DataFrame, participant_id: str) -> pd.DataFrame:
        """
        Finds all rows matching a given participant ID (case-insensitive).
        
        :param df: The DataFrame to search.
        :param participant_id: The ID to find.
        :return: A DataFrame containing all matching rows (empty if none).
        """
        participant_id_clean = participant_id.strip().lower()
        if participant_id_clean == 'nan':
            return df[df['ID'].isna()]
            
        df_id_as_str = df['ID'].astype(str).str.strip().str.lower()
        existing_rows = df[df_id_as_str == participant_id_clean]
        return existing_rows

    def save_data(self, df: pd.DataFrame, participant_id: Optional[str] = None, *, archive_previous: bool = True):
        """
        Saves the DataFrame back to the Excel file.
        Adds/Updates 'session date' for the specific participant row if ID is provided.
        Optionally archives the existing file before overwriting.
        
        PRESERVES FORMATTING: Uses openpyxl to write values into existing cells 
        without destroying cell styles (colors).
        
        :param df: The DataFrame to save.
        :param participant_id: The specific participant ID to stamp with the current date.
                               If None, no session date update occurs.
        :param archive_previous: If True (default), copy the existing file
                                 to 'archive_conditions_file/' with a timestamp
                                 before saving.
        :raises RepositoryError: If permission is denied, write fails, or archiving fails.
        """
        # Ensure self.excel_file_path is a Path object for manipulation
        file_path = Path(self.excel_file_path)

        # 1. Archive logic
        if archive_previous and file_path.exists():
            try:
                # Define archive directory (relative to the excel file)
                archive_dir = file_path.parent / "archive_conditions_file"
                
                # Create directory if it doesn't exist
                archive_dir.mkdir(parents=True, exist_ok=True)
                
                # Generate timestamp string
                now = datetime.now()
                # MODIFICATION: Added seconds to timestamp format
                timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
                
                # Create new archive filename
                file_stem = file_path.stem
                file_suffix = file_path.suffix
                archive_file_name = f"{file_stem}_{timestamp}{file_suffix}"
                archive_path = archive_dir / archive_file_name
                
                # Copy the existing file to the archive location
                shutil.copy2(file_path, archive_path)
                
            except Exception as e:
                # Catch any archiving error (permissions, disk space, etc.)
                raise RepositoryError(f"Failed to archive existing file: {e}")

        # 2. Update Logic (In-Memory DataFrame)
        try:
            # MODIFICATION: Update 'session date' only for the specific participant_id row
            if participant_id is not None:
                now_date_str = datetime.now().strftime("%Y-%m-%d")
                
                # Ensure 'session date' column exists
                if 'session date' not in df.columns:
                    df['session date'] = None

                # Locate the row(s) matching the ID (case-insensitive handling)
                participant_id_clean = participant_id.strip().lower()
                
                # Create a boolean mask for the matching ID
                mask = df['ID'].astype(str).str.strip().str.lower() == participant_id_clean
                
                # Apply the date only to matching rows
                if mask.any():
                    df.loc[mask, 'session date'] = now_date_str
            
            # 3. Save Logic (Preserving Colors using openpyxl)
            # Load the existing workbook to preserve styles
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            except FileNotFoundError:
                # If file doesn't exist, fallback to standard pandas write (no colors to preserve)
                df.to_excel(file_path, index=False)
                return

            # Write DataFrame headers to ensure alignment (Row 1)
            for c_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=c_idx)
                cell.value = col_name
                # Note: We do not reset cell.style here

            # Write DataFrame data (Rows 2+)
            # enumerate(..., 2) starts at row 2 (Excel is 1-indexed, 1 is header)
            for r_idx, row in enumerate(df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    # Handle NaN/None values for Excel compatibility
                    if pd.isna(value):
                        value = None
                    
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    # Crucial: We are modifying .value, but NOT .fill, .font, or .border

            wb.save(file_path)

        except PermissionError:
            raise RepositoryError(f"Permission denied. Cannot write to {file_path}.")
        except Exception as e:
            raise RepositoryError(f"Error writing to Excel file: {e}")